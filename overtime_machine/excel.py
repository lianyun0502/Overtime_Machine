import openpyxl
from itertools import islice
from datetime import datetime, timedelta, time
import pandas as pd

def get_exel_data(file:str, e_number:str)->list:
    wb = openpyxl.load_workbook(file)
    print(wb.sheetnames)
    my_data=[]
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f'sheet title :{sheet.title}')
        # print(sheet.max_row)
        # print(sheet.max_column)
        # print(sheet.tables)
        data = sheet.values
        cols = next(data)[1:]
        data2 = list(data)
        for rowin in data2:
            if e_number.upper() == str(rowin[4]).upper() :
                if type(rowin[0]) is str:
                    rowin = list(rowin)
                    rowin[0] = datetime.strptime(rowin[0], r'%m/%d/%Y')
                    rowin = tuple(rowin)
                my_data.append(rowin)
        print('-'*30)
    wb.close()
    return my_data

def get_start_overtime(time_:datetime, weekday:str, work_hour:int=9, start_time:time=time(hour=19), )->time:
    if weekday in ('星期六', '星期日'):
        start_time = time_
        return start_time
    
    time_ = time_.replace(hour=time_.hour + work_hour)
    if time_ > start_time:
        start_time = time_

    return start_time

def get_overtime(start_time:time, end_time:time)->float:
    if end_time < start_time:
        over_time = time(hour=0)
    else:
        over_time = timedelta(hours=end_time.hour, minutes=end_time.minute) - timedelta(hours=start_time.hour, minutes=start_time.minute)
        over_time = time(hour=over_time.seconds//3600, minute=(over_time.seconds//60)%60)

    if over_time.minute < 20:
        # over_time = over_time.replace(minute=0)
        over_time = over_time.hour
    elif 50 > over_time.minute >= 20:
        # over_time = over_time.replace(minute=30)
        over_time = over_time.hour + 0.5
    else :
        # over_time = over_time.replace(hour=over_time.hour+1, minute=0)
        over_time = over_time.hour + 1
    return over_time

def get_DataFrame(data, date:datetime)->pd.DataFrame:
    # print(my_data)
    columns = ('Date', 
            'Weekday', 
            'Department', 
            'Employee Name', 
            'Employee No.', 
            'First Record', 
            'Last Record', 
            'Working hours', 
            'Start Over Time', 
            'Over Time Hours',)
    df = pd.DataFrame(data, columns=columns)

    df = df[df['Date'] >= date]

    df['Start Over Time'] = df.apply(lambda x: get_start_overtime(x['First Record'], x['Weekday']), axis=1)
    df['Over Time Hours'] = df.apply(lambda x: get_overtime(x['Start Over Time'], x['Last Record']), axis=1)
    df.sort_values(by=['Over Time Hours','Date'], inplace=True)
    pd.set_option('display.unicode.ambiguous_as_wide', True)
    pd.set_option('display.unicode.east_asian_width', True)
    print(f'Total over time:{df["Over Time Hours"].sum()}')
    df = df.iloc[-14:]
    df.sort_values(by=['Date'], inplace=True)

    print(df)
    return df

if __name__ == '__main__':
    name = 'Eric Li'
    date = datetime(2023, 8, 21)
    my_data = get_exel_data(r'20230821~20230920 Attendance records.xlsx', name)
    df = get_DataFrame(my_data, date)
    print(df)